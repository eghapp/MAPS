#!/usr/bin/env python3
"""
MAPS Board Reader (Inference)
==============================
Reads a Crossplay game screenshot and outputs the 15x15 board state
plus the tray tiles.

Uses trained models:
  - board_classifier_v3.pkl to classify each cell as T (tile), B (bonus), or E (empty)
  - tray_classifier_v2.pkl to read tray tile letters
  - Tesseract OCR as fallback for board tile letters

Inputs:
  - A Crossplay game screenshot (.png or .jpeg)
  - The trained model files
  - The workbook with Blank Board sheet (for bonus square labels)

Outputs:
  - 15x15 board grid with letters, bonus labels, and dots for empty cells
  - Tray contents (up to 7 tiles)

Usage:
  python maps_board_reader.py --image game.png \\
      --board-model board_classifier_v3.pkl \\
      --tray-model tray_classifier_v2.pkl \\
      --workbook Cross_Play_Games_With_Images_v21.xlsx

Requirements:
  pip install openpyxl scikit-learn numpy Pillow pytesseract
  apt install tesseract-ocr  (for OCR letter reading)

Changes in v3 (from v2):
  - Cell images resized to 60x60 before feature extraction (matches v3 training)
  - Integrated tray reading via tray_classifier_v2.pkl
  - Handles variable tray sizes (1-7 tiles) and end-game (no tray)
  - Compatible with board_classifier_v3.pkl (14 features, cell_resize=(60,60))

Author: Claude (for Ed)
Date: 2026-04-07
Version: 3.0
"""

import argparse
import pickle
import sys

import numpy as np
from PIL import Image, ImageFilter, ImageOps


# ============================================================
# Feature extraction (must match training exactly)
# ============================================================

FEATURE_NAMES = [
    'gray_mean', 'gray_std', 'dark_pixels_pct', 'very_dark_pixels_pct',
    'edge_strength', 'high_edge_count', 'contrast',
    'r_mean', 'g_mean', 'b_mean', 'color_variance',
    'uniformity', 'center_darkness', 'center_dark_pixels'
]

BOARD_CELL_SIZE = (60, 60)   # Must match board_classifier_v3 training
TRAY_TILE_SIZE = (80, 80)    # Must match tray_classifier_v2 training


def extract_board_features(cell_img):
    """Extract 14 visual features from a board cell image, resized to standard size."""
    cell_img = cell_img.resize(BOARD_CELL_SIZE, Image.LANCZOS)
    return _extract_features_core(cell_img)


def extract_tray_features(tile_img):
    """Extract 14 visual features from a tray tile image, resized to standard size."""
    tile_img = tile_img.resize(TRAY_TILE_SIZE)
    return _extract_features_core(tile_img)


def _extract_features_core(cell_img):
    """Core 14-feature extraction (shared by board and tray)."""
    gray = ImageOps.grayscale(cell_img)
    ga = np.array(gray, dtype=float)
    rgb = np.array(cell_img.convert('RGB'), dtype=float)
    h, w = ga.shape
    if h == 0 or w == 0:
        return [0] * 14

    edges = np.array(gray.filter(ImageFilter.FIND_EDGES), dtype=float)

    ch, cw = max(1, h // 4), max(1, w // 4)
    center = ga[ch:h - ch, cw:w - cw]
    if center.size == 0:
        center = ga

    total_pixels = ga.size
    hist, _ = np.histogram(ga.flatten(), bins=256, range=(0, 256))
    p = hist / total_pixels
    uniformity = float(np.sum(p * p))

    return [
        ga.mean(), ga.std(),
        (ga < 128).sum() / total_pixels * 100,
        (ga < 64).sum() / total_pixels * 100,
        edges.mean(), float((edges > 128).sum()),
        float(ga.max() - ga.min()),
        rgb[:, :, 0].mean(), rgb[:, :, 1].mean(), rgb[:, :, 2].mean(),
        rgb.var(), uniformity,
        center.mean(), float((center < 128).sum()),
    ]


# ============================================================
# Board detection
# ============================================================

def find_board(img):
    """
    Auto-detect the 15x15 board grid region in a Crossplay screenshot.
    Board width ~ 95% of image width, board is square, starts after header.
    Returns: (x1, y1, x2, y2)
    """
    arr = np.array(img.convert('RGB'))
    h, w = arr.shape[:2]
    gray = np.mean(arr, axis=2)

    board_w = int(w * 0.95)
    x1 = (w - board_w) // 2
    x2 = x1 + board_w
    side = board_w

    row_frac = [(gray[y, x1:x2] < 240).mean() for y in range(h)]

    board_top = int(h * 0.25)
    for y in range(int(h * 0.10), int(h * 0.40)):
        if row_frac[y] > 0.50:
            if all(row_frac[min(y + i, h - 1)] > 0.40 for i in range(10)):
                board_top = y
                break

    if board_top + side > h:
        side = h - board_top

    cx = w // 2
    x1 = max(0, cx - side // 2)
    x2 = x1 + side

    return x1, board_top, x2, board_top + side


# ============================================================
# Tray detection
# ============================================================

def find_tray(img):
    """
    Find the tray region at the bottom of a Crossplay screenshot.
    Returns: (tray_top, tray_bottom) or None if no tray found.
    """
    arr = np.array(img.convert('RGB'))
    h, w = arr.shape[:2]

    blue_frac = [
        ((arr[y, :, 0] > 30) & (arr[y, :, 0] < 120) & (arr[y, :, 2] > 140)).mean()
        for y in range(h)
    ]

    bands = []
    in_band = False
    band_start = 0
    for y in range(h):
        if blue_frac[y] > 0.30:
            if not in_band:
                band_start = y
                in_band = True
        else:
            if in_band:
                if y - band_start > 15:
                    bands.append((band_start, y, y - band_start))
                in_band = False
    if in_band and h - band_start > 15:
        bands.append((band_start, h, h - band_start))

    if not bands:
        return None

    tray_top, tray_bottom, tray_h = bands[-1]
    if tray_top < h * 0.60:
        return None
    if tray_h > 250:
        return None

    return tray_top, tray_bottom


def extract_tray_tiles(img, tray_top, tray_bottom):
    """Split the tray region into 7 individual tile images."""
    w = img.size[0]
    tile_w = w / 7
    tiles = []
    for i in range(7):
        x1 = int(i * tile_w)
        x2 = int((i + 1) * tile_w)
        tile = img.crop((x1, tray_top, x2, tray_bottom))
        tiles.append(tile)
    return tiles


# ============================================================
# OCR for tile letter reading (fallback)
# ============================================================

def ocr_tile(cell_img):
    """
    Read the letter from a blue tile cell using Tesseract OCR.
    Returns single uppercase letter or '?' on failure.
    """
    try:
        import pytesseract
    except ImportError:
        return '?'

    arr = np.array(cell_img.convert('RGB'))
    h, w = arr.shape[:2]

    mx, my = int(w * 0.18), int(h * 0.12)
    inner = arr[my:h - my, mx:w - mx]

    r_ch, g_ch, b_ch = inner[:, :, 0], inner[:, :, 1], inner[:, :, 2]
    white_mask = (r_ch > 180) & (g_ch > 180) & (b_ch > 180)

    binary = np.where(white_mask, 0, 255).astype(np.uint8)
    bin_img = Image.fromarray(binary)

    big = bin_img.resize((bin_img.width * 5, bin_img.height * 5), Image.NEAREST)
    big = ImageOps.expand(big, border=20, fill=255)

    text = pytesseract.image_to_string(
        big, config='--psm 10 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    ).strip()
    if len(text) == 1 and text.isalpha():
        return text.upper()

    text = pytesseract.image_to_string(
        big, config='--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    ).strip()
    if len(text) >= 1 and text[0].isalpha():
        return text[0].upper()

    return '?'


# ============================================================
# Bonus square layout
# ============================================================

def load_bonus_squares(workbook_path=None):
    """Load bonus square labels from the Blank Board sheet."""
    STANDARD = [
        ['3L', '',   '',   '3W', '',   '',   '',   '2L', '',   '',   '',   '3W', '',   '',   '3L'],
        ['',   '2W', '',   '',   '',   '3L', '',   '',   '',   '3L', '',   '',   '',   '2W', ''  ],
        ['',   '',   '',   '',   '2L', '',   '',   '',   '',   '',   '2L', '',   '',   '',   ''  ],
        ['3W', '',   '',   '2L', '',   '',   '',   '2W', '',   '',   '',   '2L', '',   '',   '3W'],
        ['',   '',   '2L', '',   '',   '',   '3L', '',   '',   '',   '3L', '',   '2L', '',   ''  ],
        ['',   '',   '',   '',   '',   '3L', '',   '',   '2L', '',   '',   '3L', '',   '',   ''  ],
        ['',   '3L', '',   '',   '',   '',   '',   '',   '',   '',   '',   '',   '',   '3L', ''  ],
        ['2L', '',   '',   '2W', '',   '2L', '',   'T',  '',   '2L', '',   '2W', '',   '',   '2L'],
        ['',   '3L', '',   '',   '',   '',   '',   '',   '',   '',   '',   '',   '',   '3L', ''  ],
        ['',   '',   '',   '',   '',   '3L', '',   '',   '2L', '',   '',   '3L', '',   '',   ''  ],
        ['',   '',   '2L', '',   '',   '',   '3L', '',   '',   '',   '3L', '',   '2L', '',   ''  ],
        ['3W', '',   '',   '2L', '',   '',   '',   '2W', '',   '',   '',   '2L', '',   '',   '3W'],
        ['',   '',   '',   '',   '2L', '',   '',   '',   '',   '',   '2L', '',   '',   '',   ''  ],
        ['',   '2W', '',   '',   '',   '3L', '',   '',   '',   '3L', '',   '',   '',   '2W', ''  ],
        ['3L', '',   '',   '3W', '',   '',   '',   '2L', '',   '',   '',   '3W', '',   '',   '3L'],
    ]

    if workbook_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(workbook_path)
            if 'Blank Board' in wb.sheetnames:
                ws = wb['Blank Board']
                bonus = [['' for _ in range(15)] for _ in range(15)]
                for r in range(1, 16):
                    for c in range(1, 16):
                        v = ws.cell(r, c).value
                        if v and str(v).strip():
                            bonus[r - 1][c - 1] = str(v).strip()
                return bonus
        except Exception:
            pass

    return STANDARD


# ============================================================
# Main inference pipeline
# ============================================================

def read_board(image_path, board_model_path, tray_model_path=None,
               workbook_path=None, do_ocr=True):
    """
    Full pipeline: load image -> detect board -> classify cells -> OCR letters
                   -> detect tray -> classify tray tiles.
    Returns: (grid, tray_letters, stats)
    """
    # Load board model
    with open(board_model_path, 'rb') as f:
        board_model = pickle.load(f)
    board_clf = board_model['classifier']
    board_scaler = board_model['scaler']

    # Load tray model (optional)
    tray_clf = None
    tray_scaler = None
    if tray_model_path:
        with open(tray_model_path, 'rb') as f:
            tray_model = pickle.load(f)
        tray_clf = tray_model['classifier']
        tray_scaler = tray_model['scaler']

    # Load bonus squares
    bonus = load_bonus_squares(workbook_path)

    # Load and process image
    img = Image.open(image_path)

    # ---- Board classification ----
    x1, y1, x2, y2 = find_board(img)
    cw = (x2 - x1) / 15
    ch = (y2 - y1) / 15

    grid = [['.' for _ in range(15)] for _ in range(15)]
    tile_count = 0
    ocr_success = 0
    ocr_fail = 0

    for r in range(15):
        for c in range(15):
            cx1 = int(x1 + c * cw)
            cy1 = int(y1 + r * ch)
            cell = img.crop((cx1, cy1, int(cx1 + cw), int(cy1 + ch)))

            f = np.array([extract_board_features(cell)])
            pred = board_clf.predict(board_scaler.transform(f))[0]

            if pred == 'T':
                tile_count += 1
                if do_ocr:
                    letter = ocr_tile(cell)
                    grid[r][c] = letter
                    if letter == '?':
                        ocr_fail += 1
                    else:
                        ocr_success += 1
                else:
                    grid[r][c] = 'T'
            elif pred == 'B' and bonus[r][c]:
                grid[r][c] = bonus[r][c]
            else:
                grid[r][c] = '.'

    # ---- Tray reading ----
    tray_letters = []
    tray_found = False
    tray_bounds = None

    result = find_tray(img)
    if result is not None:
        tray_top, tray_bottom = result
        tray_found = True
        tray_bounds = (tray_top, tray_bottom)
        tiles = extract_tray_tiles(img, tray_top, tray_bottom)

        for tile in tiles:
            if tray_clf is not None:
                f = np.array([extract_tray_features(tile)])
                letter = tray_clf.predict(tray_scaler.transform(f))[0]
                tray_letters.append(letter)
            else:
                tray_letters.append('?')

    stats = {
        'image_size': img.size,
        'board_bounds': (x1, y1, x2, y2),
        'cell_size': (int(cw), int(ch)),
        'tiles': tile_count,
        'ocr_success': ocr_success,
        'ocr_fail': ocr_fail,
        'tray_found': tray_found,
        'tray_bounds': tray_bounds,
        'tray_tiles': len(tray_letters),
    }

    return grid, tray_letters, stats


def print_grid(grid, tray_letters, stats):
    """Print the board grid and tray in human-readable format."""
    print(f"Image: {stats['image_size'][0]}x{stats['image_size'][1]}  |  "
          f"Board: {stats['board_bounds']}  |  Cell: {stats['cell_size'][0]}x{stats['cell_size'][1]}px")
    print(f"Tiles: {stats['tiles']}  |  OCR: {stats['ocr_success']}/{stats['tiles']} "
          f"({stats['ocr_fail']} failed)")
    print()
    print("     A  B  C  D  E  F  G  H  I  J  K  L  M  N  O")
    for r in range(15):
        row_str = f"{r + 1:2d}  "
        for c in range(15):
            cell = grid[r][c]
            row_str += f"{cell:>2} "
        print(row_str)

    print()
    if tray_letters:
        print(f"Tray ({stats['tray_tiles']} tiles): {' '.join(tray_letters)}")
        if stats['tray_bounds']:
            print(f"  Tray region: y={stats['tray_bounds'][0]}-{stats['tray_bounds'][1]}")
    else:
        print("Tray: not found (end-game or detection failed)")


def main():
    parser = argparse.ArgumentParser(
        description='MAPS Board Reader v3 - reads Crossplay screenshots using trained ML models'
    )
    parser.add_argument('--image', required=True,
                        help='Path to Crossplay game screenshot')
    parser.add_argument('--board-model', default='board_classifier_v3.pkl',
                        help='Path to board classifier .pkl file')
    parser.add_argument('--tray-model', default='tray_classifier_v2.pkl',
                        help='Path to tray classifier .pkl file (optional)')
    parser.add_argument('--workbook', default=None,
                        help='Path to workbook with Blank Board sheet')
    parser.add_argument('--no-ocr', action='store_true',
                        help='Skip OCR, just show T/B/E classification')
    # Backward compatibility alias
    parser.add_argument('--model', default=None,
                        help='(deprecated) Alias for --board-model')

    args = parser.parse_args()

    board_model = args.board_model
    if args.model:
        board_model = args.model

    grid, tray_letters, stats = read_board(
        args.image, board_model, args.tray_model,
        args.workbook, do_ocr=not args.no_ocr
    )
    print_grid(grid, tray_letters, stats)


if __name__ == '__main__':
    main()
