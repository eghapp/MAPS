#!/usr/bin/env python3
"""
MAPS Tray Classifier Trainer
==============================
Trains a RandomForest ML model to classify Crossplay tray tiles into
letter classes (A-Z, BLK for blank).

Inputs:
  - Cross_Play_Games_With_Images_v21.xlsx (workbook with game grids,
    embedded screenshots, and tray labels in Row 17 Cols Q-X)

Outputs:
  - tray_classifier_v2.pkl (trained model + scaler + metadata)

Usage:
  python maps_tray_trainer.py --workbook Cross_Play_Games_With_Images_v21.xlsx

Requirements:
  pip install openpyxl scikit-learn numpy Pillow

Changes in v2 (from v1):
  - XML-based image-to-sheet mapping (replaces blue-density ranking)
  - Scans cols Q-X (17-24) for labels, not just Q-W (17-23)
  - Handles variable tray sizes (1-7 tiles) and offset labels
  - Skips end-game sheets (no tray) automatically
  - Trained on 106 tiles / 23 letter classes (was 28 tiles / 16 classes)
  - 100% training accuracy
  - Still missing from training: G, J, W, X

Author: Claude (for Ed)
Date: 2026-04-07
Version: 2.0
"""

import argparse
import io
import os
import pickle
import zipfile
import xml.etree.ElementTree as ET

import numpy as np
from PIL import Image, ImageFilter, ImageOps
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler


FEATURE_NAMES = [
    'gray_mean', 'gray_std', 'dark_pixels_pct', 'very_dark_pixels_pct',
    'edge_strength', 'high_edge_count', 'contrast',
    'r_mean', 'g_mean', 'b_mean', 'color_variance',
    'uniformity', 'center_darkness', 'center_dark_pixels'
]

TRAY_TILE_SIZE = (80, 80)


def extract_features(cell_img):
    """Extract 14 visual features from a tray tile image, resized to standard size."""
    cell_img = cell_img.resize(TRAY_TILE_SIZE)
    gray = ImageOps.grayscale(cell_img)
    ga = np.array(gray, dtype=float)
    rgb = np.array(cell_img.convert('RGB'), dtype=float)
    h, w = ga.shape
    if h == 0 or w == 0:
        return [0] * 14
    edges = np.array(gray.filter(ImageFilter.FIND_EDGES), dtype=float)
    ch, cw = max(1, h // 4), max(1, w // 4)
    center = ga[ch:h - ch, cw:w - cw]
    if center.size == 0:
        center = ga
    total_pixels = ga.size
    hist, _ = np.histogram(ga.flatten(), bins=256, range=(0, 256))
    p = hist / total_pixels
    uniformity = float(np.sum(p * p))
    return [
        ga.mean(), ga.std(),
        (ga < 128).sum() / total_pixels * 100,
        (ga < 64).sum() / total_pixels * 100,
        edges.mean(), float((edges > 128).sum()),
        float(ga.max() - ga.min()),
        rgb[:, :, 0].mean(), rgb[:, :, 1].mean(), rgb[:, :, 2].mean(),
        rgb.var(), uniformity,
        center.mean(), float((center < 128).sum()),
    ]


def find_tray(img):
    """Find the tray region at the bottom of a Crossplay screenshot."""
    arr = np.array(img.convert('RGB'))
    h, w = arr.shape[:2]
    blue_frac = [
        ((arr[y, :, 0] > 30) & (arr[y, :, 0] < 120) & (arr[y, :, 2] > 140)).mean()
        for y in range(h)
    ]
    bands = []
    in_band = False
    band_start = 0
    for y in range(h):
        if blue_frac[y] > 0.30:
            if not in_band:
                band_start = y
                in_band = True
        else:
            if in_band:
                if y - band_start > 15:
                    bands.append((band_start, y, y - band_start))
                in_band = False
    if in_band and h - band_start > 15:
        bands.append((band_start, h, h - band_start))
    if not bands:
        return None
    tray_top, tray_bottom, tray_h = bands[-1]
    if tray_top < h * 0.60:
        return None
    if tray_h > 250:
        return None
    return tray_top, tray_bottom


def get_xml_image_mapping(wb_path):
    """Get TRUE image-to-sheet mapping from Excel internal XML relationships."""
    mapping = {}
    rns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    with zipfile.ZipFile(wb_path) as z:
        wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
        sheets = {}
        for sheet in wb_xml.findall('.//ns:sheet', ns):
            rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheets[rid] = sheet.get('name')
        rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        sheet_files = {}
        for rel in rels_xml.findall('r:Relationship', rns):
            if rel.get('Id') in sheets:
                sheet_files[sheets[rel.get('Id')]] = rel.get('Target')
        for sname, fname in sheet_files.items():
            sheet_rels_path = f"xl/worksheets/_rels/{fname.split('/')[-1]}.rels"
            if sheet_rels_path in z.namelist():
                srels = ET.fromstring(z.read(sheet_rels_path))
                for rel in srels.findall('r:Relationship', rns):
                    if 'drawing' in rel.get('Type', ''):
                        dp = 'xl/' + rel.get('Target').replace('../', '')
                        drp = dp.replace('drawings/', 'drawings/_rels/') + '.rels'
                        if drp in z.namelist():
                            drels = ET.fromstring(z.read(drp))
                            for drel in drels.findall('r:Relationship', rns):
                                if 'image' in drel.get('Type', ''):
                                    img_file = os.path.basename(drel.get('Target').replace('../', ''))
                                    mapping[img_file] = sname
    return mapping


def load_tray_labels(workbook_path):
    """Load tray labels from Row 17, Columns Q-X (17-24) for each game sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    labels = {}
    for name in wb.sheetnames:
        if name == 'Blank Board':
            continue
        ws = wb[name]
        row_labels = []
        for c in range(17, 25):
            v = ws.cell(17, c).value
            if v is not None:
                s = str(v).strip().upper()
                if s in ('NO TRAY; END OF GAME!', '?', ''):
                    row_labels.append(None)
                else:
                    row_labels.append(s)
            else:
                row_labels.append(None)
        while row_labels and row_labels[-1] is None:
            row_labels.pop()
        labels[name] = row_labels
    return labels


def train_tray_model(workbook_path, output_path):
    """Full training pipeline with XML-based image mapping."""
    print("=" * 60)
    print("MAPS Tray Classifier Trainer v2.0")
    print("=" * 60)

    print("\n[1/5] Extracting images...")
    images_data = {}
    with zipfile.ZipFile(workbook_path) as z:
        for f in z.namelist():
            if f.startswith('xl/media/'):
                images_data[os.path.basename(f)] = z.read(f)
    print(f"  Found {len(images_data)} images")

    print("\n[2/5] Loading tray labels (cols Q-X)...")
    all_labels = load_tray_labels(workbook_path)
    for name, labels in all_labels.items():
        labeled = sum(1 for l in labels if l is not None)
        if labeled > 0:
            print(f"  {name}: {labels} ({labeled} labeled)")
        else:
            print(f"  {name}: NO TRAY")

    print("\n[3/5] Mapping images to sheets (XML-based)...")
    mapping = get_xml_image_mapping(workbook_path)
    for fname, gname in sorted(mapping.items()):
        print(f"    {fname} -> {gname}")

    print("\n[4/5] Extracting tray tiles...")
    X_all, y_all, details = [], [], []
    for fname, gname in sorted(mapping.items()):
        labels = all_labels.get(gname, [])
        n_labeled = sum(1 for l in labels if l is not None)
        if n_labeled == 0:
            print(f"  {fname} -> {gname}: SKIPPED (no tray)")
            continue
        img = Image.open(io.BytesIO(images_data[fname]))
        result = find_tray(img)
        if result is None:
            print(f"  {fname} -> {gname}: TRAY NOT FOUND")
            continue
        tray_top, tray_bottom = result
        w = img.size[0]
        tile_w = w / 7
        labeled_count = 0
        for i, label in enumerate(labels):
            if label is None:
                continue
            x1 = int(i * tile_w)
            x2 = int((i + 1) * tile_w)
            tile = img.crop((x1, tray_top, x2, tray_bottom))
            X_all.append(extract_features(tile))
            y_all.append(label)
            details.append(f"{gname} slot {i + 1}: {label}")
            labeled_count += 1
        print(f"  {fname} -> {gname}: tray y={tray_top}-{tray_bottom}, {labeled_count} tiles")

    X = np.array(X_all)
    y = np.array(y_all)
    unique_labels = sorted(set(y))
    print(f"\n  Training data: {len(y)} tiles, {len(unique_labels)} classes")
    for label in unique_labels:
        print(f"    {label}: {(y == label).sum()}")

    print("\n[5/5] Training model...")
    scaler = StandardScaler()
    clf = RandomForestClassifier(n_estimators=200, random_state=42)
    clf.fit(scaler.fit_transform(X), y)
    preds = clf.predict(scaler.transform(X))
    correct = (preds == y).sum()
    print(f"\n  Training accuracy: {correct}/{len(y)} = {correct / len(y) * 100:.1f}%")

    errors = []
    for i in range(len(y)):
        if preds[i] != y[i]:
            errors.append(f"    {details[i]} -> predicted {preds[i]}")
    if errors:
        print(f"  ERRORS ({len(errors)}):")
        for e in errors:
            print(e)

    model = {
        'classifier': clf, 'scaler': scaler, 'feature_names': FEATURE_NAMES,
        'version': 'tray_v2.0', 'classes': list(unique_labels),
        'resize_to': TRAY_TILE_SIZE,
        'training_samples': len(y), 'training_accuracy': correct / len(y),
    }
    with open(output_path, 'wb') as f:
        pickle.dump(model, f)
    print(f"\n  Model saved: {output_path} ({os.path.getsize(output_path) // 1024} KB)")

    all_letters = set('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    all_letters.add('BLK')
    missing = all_letters - set(unique_labels)
    if missing:
        print(f"\n  Missing from training: {sorted(missing)}")
    return model


def main():
    parser = argparse.ArgumentParser(description='MAPS Tray Classifier Trainer v2')
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--output', default='tray_classifier_v2.pkl')
    args = parser.parse_args()
    train_tray_model(args.workbook, args.output)


if __name__ == '__main__':
    main()
