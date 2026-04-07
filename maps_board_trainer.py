#!/usr/bin/env python3
"""
MAPS Board Classifier Trainer
==============================
Trains a RandomForest ML model to classify Crossplay board cells as:
  T = Tile (letter placed)
  B = Bonus square (uncovered 2L/3L/2W/3W)
  E = Empty

Inputs:
  - Cross_Play_Games_With_Images_v21.xlsx (workbook with 18 game grids + embedded screenshots)

Outputs:
  - board_classifier_v3.pkl (trained model + scaler + feature names)

Usage:
  python maps_board_trainer.py --workbook Cross_Play_Games_With_Images_v21.xlsx

Requirements:
  pip install openpyxl scikit-learn numpy Pillow

Changes in v3 (from v2):
  - XML-based image-to-sheet mapping (replaces blue-density ranking)
  - Cell images resized to 60x60 before feature extraction (resolution-invariant)
  - Trained on 18 images / 4050 cells (was 5 images / 1125 cells)
  - 96% training accuracy across all 18 game images

Author: Claude (for Ed)
Date: 2026-04-07
Version: 3.0
"""

import argparse
import os
import pickle
import zipfile
import xml.etree.ElementTree as ET

import numpy as np
from PIL import Image, ImageFilter, ImageOps
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler


FEATURE_NAMES = [
    'gray_mean', 'gray_std', 'dark_pixels_pct', 'very_dark_pixels_pct',
    'edge_strength', 'high_edge_count', 'contrast',
    'r_mean', 'g_mean', 'b_mean', 'color_variance',
    'uniformity', 'center_darkness', 'center_dark_pixels'
]

CELL_SIZE = (60, 60)


def extract_features(cell_img):
    """Extract 14 visual features from a board cell image, resized to standard size."""
    cell_img = cell_img.resize(CELL_SIZE, Image.LANCZOS)
    gray = ImageOps.grayscale(cell_img)
    ga = np.array(gray, dtype=float)
    rgb = np.array(cell_img.convert('RGB'), dtype=float)
    h, w = ga.shape
    if h == 0 or w == 0:
        return [0] * 14
    edges = np.array(gray.filter(ImageFilter.FIND_EDGES), dtype=float)
    ch, cw = max(1, h // 4), max(1, w // 4)
    center = ga[ch:h - ch, cw:w - cw]
    if center.size == 0:
        center = ga
    total_pixels = ga.size
    hist, _ = np.histogram(ga.flatten(), bins=256, range=(0, 256))
    p = hist / total_pixels
    uniformity = float(np.sum(p * p))
    return [
        ga.mean(), ga.std(),
        (ga < 128).sum() / total_pixels * 100,
        (ga < 64).sum() / total_pixels * 100,
        edges.mean(), float((edges > 128).sum()),
        float(ga.max() - ga.min()),
        rgb[:, :, 0].mean(), rgb[:, :, 1].mean(), rgb[:, :, 2].mean(),
        rgb.var(), uniformity,
        center.mean(), float((center < 128).sum()),
    ]


def find_board(img):
    """Auto-detect the 15x15 board region. Board width ~ 95% of image, square."""
    arr = np.array(img.convert('RGB'))
    h, w = arr.shape[:2]
    gray = np.mean(arr, axis=2)
    board_w = int(w * 0.95)
    x1 = (w - board_w) // 2
    x2 = x1 + board_w
    side = board_w
    row_frac = [(gray[y, x1:x2] < 240).mean() for y in range(h)]
    board_top = int(h * 0.25)
    for y in range(int(h * 0.10), int(h * 0.40)):
        if row_frac[y] > 0.50:
            if all(row_frac[min(y + i, h - 1)] > 0.40 for i in range(10)):
                board_top = y
                break
    if board_top + side > h:
        side = h - board_top
    cx = w // 2
    x1 = max(0, cx - side // 2)
    x2 = x1 + side
    return x1, board_top, x2, board_top + side


def get_xml_image_mapping(wb_path):
    """Get TRUE image-to-sheet mapping from Excel internal XML relationships."""
    mapping = {}
    rns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    with zipfile.ZipFile(wb_path) as z:
        wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
        sheets = {}
        for sheet in wb_xml.findall('.//ns:sheet', ns):
            rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheets[rid] = sheet.get('name')
        rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        sheet_files = {}
        for rel in rels_xml.findall('r:Relationship', rns):
            if rel.get('Id') in sheets:
                sheet_files[sheets[rel.get('Id')]] = rel.get('Target')
        for sname, fname in sheet_files.items():
            sheet_rels_path = f"xl/worksheets/_rels/{fname.split('/')[-1]}.rels"
            if sheet_rels_path in z.namelist():
                srels = ET.fromstring(z.read(sheet_rels_path))
                for rel in srels.findall('r:Relationship', rns):
                    if 'drawing' in rel.get('Type', ''):
                        drawing_path = 'xl/' + rel.get('Target').replace('../', '')
                        draw_rels_path = drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
                        if draw_rels_path in z.namelist():
                            drels = ET.fromstring(z.read(draw_rels_path))
                            for drel in drels.findall('r:Relationship', rns):
                                if 'image' in drel.get('Type', ''):
                                    img_file = os.path.basename(drel.get('Target').replace('../', ''))
                                    mapping[img_file] = sname
    return mapping


def load_grids(workbook_path):
    """Load 15x15 grids labeled T/B/E from the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path)
    prems = {'3W', '2W', '3L', '2L', 'T'}
    blank = [['E'] * 15 for _ in range(15)]
    if 'Blank Board' in wb.sheetnames:
        ws0 = wb['Blank Board']
        for r in range(1, 16):
            for c in range(1, 16):
                v = ws0.cell(r, c).value
                if v and str(v).strip() in prems:
                    blank[r - 1][c - 1] = 'B'
    grids = {}
    for name in wb.sheetnames:
        if name == 'Blank Board':
            continue
        ws = wb[name]
        grid = [row[:] for row in blank]
        for r in range(1, 16):
            for c in range(1, 16):
                v = ws.cell(r, c).value
                s = str(v).strip() if v else ''
                if s and s not in prems and len(s) == 1 and s.isalpha():
                    grid[r - 1][c - 1] = 'T'
        tiles = sum(1 for r in range(15) for c in range(15) if grid[r][c] == 'T')
        grids[name] = grid
        print(f"  Grid '{name}': {tiles} tiles")
    return grids


def train_model(wb_path, output_path):
    """Full training pipeline using XML-based image mapping."""
    print("=" * 60)
    print("MAPS Board Classifier Trainer v3.0")
    print("=" * 60)

    print("\n[1/4] Loading grids from workbook...")
    grids = load_grids(wb_path)

    print("\n[2/4] Extracting images...")
    images = {}
    with zipfile.ZipFile(wb_path, 'r') as z:
        for f in sorted(z.namelist()):
            if f.startswith('xl/media/'):
                images[os.path.basename(f)] = z.read(f)
    print(f"  Found {len(images)} images")

    print("\n[3/4] Mapping images to grids (XML-based)...")
    mapping = get_xml_image_mapping(wb_path)
    for fname, gname in sorted(mapping.items()):
        tiles = sum(1 for r in range(15) for c in range(15) if grids[gname][r][c] == 'T')
        print(f"    {fname} -> {gname} ({tiles} tiles)")

    print("\n[4/4] Training model...")
    X_all, y_all = [], []
    import io
    for fname, gname in sorted(mapping.items()):
        img = Image.open(io.BytesIO(images[fname]))
        x1, y1, x2, y2 = find_board(img)
        cw = (x2 - x1) / 15
        ch = (y2 - y1) / 15
        grid = grids[gname]
        for r in range(15):
            for c in range(15):
                cell = img.crop((int(x1 + c * cw), int(y1 + r * ch),
                                 int(x1 + (c + 1) * cw), int(y1 + (r + 1) * ch)))
                X_all.append(extract_features(cell))
                y_all.append(grid[r][c])
        print(f"  {fname}: board=({x1},{y1})-({x2},{y2}), cell={cw:.0f}x{ch:.0f}px")

    X = np.array(X_all)
    y = np.array(y_all)
    print(f"\n  Training data: {X.shape[0]} cells")
    print(f"    T={( y == 'T').sum()}, B={(y == 'B').sum()}, E={(y == 'E').sum()}")

    scaler = StandardScaler()
    clf = RandomForestClassifier(n_estimators=200, random_state=42)
    clf.fit(scaler.fit_transform(X), y)

    preds = clf.predict(scaler.transform(X))
    acc = (preds == y).sum() / len(y)
    print(f"\n  Training accuracy: {(preds == y).sum()}/{len(y)} = {acc * 100:.2f}%")

    model = {
        'classifier': clf, 'scaler': scaler, 'feature_names': FEATURE_NAMES,
        'version': 'v3.0', 'cell_resize': CELL_SIZE,
        'training_images': len(mapping), 'training_cells': len(y),
    }
    with open(output_path, 'wb') as f:
        pickle.dump(model, f)
    print(f"  Model saved: {output_path} ({os.path.getsize(output_path) // 1024} KB)")
    return model


def main():
    parser = argparse.ArgumentParser(description='MAPS Board Classifier Trainer v3')
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--output', default='board_classifier_v3.pkl')
    args = parser.parse_args()
    train_model(args.workbook, args.output)


if __name__ == '__main__':
    main()
